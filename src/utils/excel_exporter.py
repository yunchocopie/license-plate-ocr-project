#!/usr/bin/env python3
"""
고급 엑셀 출력 기능 모듈

이미지 포함, 통계 정보, 다양한 포맷을 지원하는 종합적인 엑셀 출력 시스템입니다.

Features:
- 이미지 포함 및 크기 조정
- 상세 통계 정보 및 차트
- 다중 시트 지원 (요약, 상세, 통계)
- 조건부 서식 및 고급 스타일링
- 배치 처리 결과 전용 포맷
- 필터링 및 정렬 기능
"""

import os
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import tempfile

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, NamedStyle,
        numbers
    )
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """고급 엑셀 출력 클래스"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl을 실행해주세요.")
            
        self.wb = None
        self.current_file = None
        self.styles_created = False
    
    def _create_styles(self):
        """엑셀 스타일 정의"""
        if self.styles_created or not self.wb:
            return
            
        # 헤더 스타일
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
        header_style.fill = PatternFill(fill_type="solid", fgColor="366092")
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 데이터 스타일
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(name="맑은 고딕", size=10)
        data_style.alignment = Alignment(horizontal='center', vertical='center')
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 통계 헤더 스타일
        stat_header_style = NamedStyle(name="stat_header_style")
        stat_header_style.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        stat_header_style.fill = PatternFill(fill_type="solid", fgColor="70AD47")
        stat_header_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # 경고 스타일 (낮은 신뢰도)
        warning_style = NamedStyle(name="warning_style")
        warning_style.font = Font(name="맑은 고딕", size=10)
        warning_style.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        warning_style.alignment = Alignment(horizontal='center', vertical='center')
        
        try:
            self.wb.add_named_style(header_style)
            self.wb.add_named_style(data_style)
            self.wb.add_named_style(stat_header_style)
            self.wb.add_named_style(warning_style)
            self.styles_created = True
        except ValueError:
            # 스타일이 이미 존재하는 경우
            pass
    
    def create_comprehensive_report(
        self, 
        results: List[Dict[str, Any]], 
        output_path: str,
        include_images: bool = True,
        include_statistics: bool = True,
        include_charts: bool = True,
        image_max_size: tuple = (120, 90)
    ) -> str:
        """종합적인 엑셀 보고서 생성
        
        Args:
            results: 처리 결과 데이터 리스트
            output_path: 출력 파일 경로
            include_images: 이미지 포함 여부
            include_statistics: 통계 시트 포함 여부
            include_charts: 차트 포함 여부
            image_max_size: 이미지 최대 크기 (width, height)
            
        Returns:
            생성된 파일 경로
        """
        
        self.wb = Workbook()
        self.current_file = output_path
        self._create_styles()
        
        # 기본 시트 제거
        self.wb.remove(self.wb.active)
        
        # 요약 시트 생성
        self._create_summary_sheet(results)
        
        # 상세 결과 시트 생성
        self._create_detailed_sheet(results, include_images, image_max_size)
        
        # 통계 시트 생성
        if include_statistics:
            self._create_statistics_sheet(results, include_charts)
        
        # 실패 분석 시트 생성 (실패한 항목이 있는 경우)
        failed_results = [r for r in results if not r.get('success', False)]
        if failed_results:
            self._create_failed_analysis_sheet(failed_results)
        
        # 파일 저장
        os.makedirs(Path(output_path).parent, exist_ok=True)
        self.wb.save(output_path)
        
        logger.info(f"종합 엑셀 보고서 생성 완료: {output_path}")
        return output_path
    
    def _create_summary_sheet(self, results: List[Dict[str, Any]]):
        """요약 시트 생성"""
        ws = self.wb.create_sheet("📊 요약", 0)
        
        # 기본 통계 계산
        total_files = len(results)
        successful_files = len([r for r in results if r.get('success', False)])
        success_rate = (successful_files / total_files * 100) if total_files > 0 else 0
        
        total_plates = sum(len(r.get('plates', [])) for r in results)
        valid_plates = sum(len([p for p in r.get('plates', []) if p.get('is_valid', False)]) for r in results)
        
        # 번호판 타입 분포
        plate_types = {}
        for result in results:
            for plate in result.get('plates', []):
                plate_type = plate.get('type', 'Unknown')
                plate_types[plate_type] = plate_types.get(plate_type, 0) + 1
        
        # 처리 시간 통계
        processing_times = [r.get('processing_time', 0) for r in results if r.get('processing_time')]
        avg_time = sum(processing_times) / len(processing_times) if processing_times else 0
        
        # 요약 정보 작성
        summary_data = [
            ["📈 처리 요약", ""],
            ["총 처리 파일", total_files],
            ["성공한 파일", successful_files],
            ["실패한 파일", total_files - successful_files],
            ["성공률", f"{success_rate:.1f}%"],
            ["", ""],
            ["🚗 번호판 정보", ""],
            ["총 감지된 번호판", total_plates],
            ["유효한 번호판", valid_plates],
            ["유효율", f"{(valid_plates / total_plates * 100):.1f}%" if total_plates > 0 else "0%"],
            ["", ""],
            ["⏱️ 성능 정보", ""],
            ["평균 처리 시간", f"{avg_time:.2f}초"],
            ["총 처리 시간", f"{sum(processing_times):.2f}초"],
            ["", ""],
            ["📋 보고서 정보", ""],
            ["생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["보고서 버전", "v2.0"]
        ]
        
        # 데이터 입력
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            
            # 섹션 헤더 스타일 적용
            if label.startswith(("📈", "🚗", "⏱️", "📋")):
                ws.cell(row=row_idx, column=1).style = "stat_header_style"
                ws.cell(row=row_idx, column=2).style = "stat_header_style"
        
        # 번호판 타입 분포 추가
        if plate_types:
            start_row = len(summary_data) + 3
            ws.cell(row=start_row, column=1, value="🏷️ 번호판 타입 분포").style = "stat_header_style"
            ws.cell(row=start_row, column=2, value="개수").style = "stat_header_style"
            
            for idx, (plate_type, count) in enumerate(sorted(plate_types.items()), start_row + 1):
                ws.cell(row=idx, column=1, value=plate_type)
                ws.cell(row=idx, column=2, value=count)
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_detailed_sheet(self, results: List[Dict[str, Any]], include_images: bool, image_size: tuple):
        """상세 결과 시트 생성"""
        ws = self.wb.create_sheet("📄 상세 결과")
        
        # 헤더 작성
        headers = ["번호", "파일명", "상태", "처리시간", "감지된 번호판 수"]
        if include_images:
            headers.insert(2, "이미지")
        
        # 번호판별 상세 정보 헤더 추가
        max_plates = max(len(r.get('plates', [])) for r in results) if results else 0
        for i in range(max_plates):
            headers.extend([
                f"번호판{i+1}_텍스트", f"번호판{i+1}_신뢰도", 
                f"번호판{i+1}_타입", f"번호판{i+1}_색상"
            ])
        
        # 헤더 스타일 적용
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.style = "header_style"
        
        # 데이터 행 작성
        current_row = 2
        for idx, result in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=result.get('file_name', ''))
            
            col_offset = 3
            
            # 이미지 추가
            if include_images:
                image_path = result.get('file_path') or result.get('image_path')
                if image_path and os.path.exists(image_path):
                    self._add_image_to_cell(ws, current_row, col_offset, image_path, image_size)
                col_offset += 1
            
            # 상태 및 기본 정보
            status = "✅ 성공" if result.get('success', False) else "❌ 실패"
            ws.cell(row=current_row, column=col_offset, value=status)
            ws.cell(row=current_row, column=col_offset + 1, value=f"{result.get('processing_time', 0):.2f}초")
            
            plates = result.get('plates', [])
            ws.cell(row=current_row, column=col_offset + 2, value=len(plates))
            
            # 번호판 상세 정보
            detail_col_start = col_offset + 3
            for plate_idx, plate in enumerate(plates):
                base_col = detail_col_start + (plate_idx * 4)
                
                ws.cell(row=current_row, column=base_col, value=plate.get('text', ''))
                
                # 신뢰도에 따른 조건부 서식
                confidence = plate.get('confidence', 0)
                conf_cell = ws.cell(row=current_row, column=base_col + 1, value=f"{confidence:.2f}")
                if confidence < 0.7:
                    conf_cell.style = "warning_style"
                
                ws.cell(row=current_row, column=base_col + 2, value=plate.get('type', ''))
                ws.cell(row=current_row, column=base_col + 3, value=plate.get('color', ''))
            
            # 행 높이 조정 (이미지가 있는 경우)
            if include_images:
                ws.row_dimensions[current_row].height = image_size[1] + 10
            
            current_row += 1
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # 이미지 열 너비 고정
        if include_images:
            ws.column_dimensions['C'].width = 20
    
    def _create_statistics_sheet(self, results: List[Dict[str, Any]], include_charts: bool):
        """통계 시트 생성"""
        ws = self.wb.create_sheet("📊 통계 분석")
        
        # 성공/실패 통계
        success_count = len([r for r in results if r.get('success', False)])
        fail_count = len(results) - success_count
        
        # 번호판 타입 통계
        plate_type_stats = {}
        confidence_stats = []
        processing_time_stats = []
        
        for result in results:
            if result.get('processing_time'):
                processing_time_stats.append(result['processing_time'])
                
            for plate in result.get('plates', []):
                # 타입 통계
                plate_type = plate.get('type', 'Unknown')
                plate_type_stats[plate_type] = plate_type_stats.get(plate_type, 0) + 1
                
                # 신뢰도 통계
                if plate.get('confidence'):
                    confidence_stats.append(plate['confidence'])
        
        # 통계 데이터 DataFrame 생성
        stats_data = []
        
        # 처리 결과 통계
        stats_data.append(["처리 결과 통계", "", ""])
        stats_data.append(["구분", "개수", "비율(%)"])
        stats_data.append(["성공", success_count, f"{success_count/len(results)*100:.1f}" if results else "0"])
        stats_data.append(["실패", fail_count, f"{fail_count/len(results)*100:.1f}" if results else "0"])
        stats_data.append(["", "", ""])
        
        # 번호판 타입 통계
        if plate_type_stats:
            stats_data.append(["번호판 타입 통계", "", ""])
            stats_data.append(["타입", "개수", "비율(%)"])
            total_plates = sum(plate_type_stats.values())
            for plate_type, count in sorted(plate_type_stats.items()):
                stats_data.append([plate_type, count, f"{count/total_plates*100:.1f}"])
            stats_data.append(["", "", ""])
        
        # 신뢰도 통계
        if confidence_stats:
            stats_data.append(["신뢰도 통계", "", ""])
            stats_data.append(["지표", "값", ""])
            stats_data.append(["평균", f"{sum(confidence_stats)/len(confidence_stats):.3f}", ""])
            stats_data.append(["최고", f"{max(confidence_stats):.3f}", ""])
            stats_data.append(["최저", f"{min(confidence_stats):.3f}", ""])
            
            # 신뢰도 구간별 통계
            high_conf = len([c for c in confidence_stats if c >= 0.9])
            med_conf = len([c for c in confidence_stats if 0.7 <= c < 0.9])
            low_conf = len([c for c in confidence_stats if c < 0.7])
            
            stats_data.append(["", "", ""])
            stats_data.append(["신뢰도 구간별 분포", "", ""])
            stats_data.append(["높음(≥0.9)", high_conf, f"{high_conf/len(confidence_stats)*100:.1f}%"])
            stats_data.append(["보통(0.7~0.9)", med_conf, f"{med_conf/len(confidence_stats)*100:.1f}%"])
            stats_data.append(["낮음(<0.7)", low_conf, f"{low_conf/len(confidence_stats)*100:.1f}%"])
            stats_data.append(["", "", ""])
        
        # 처리 시간 통계
        if processing_time_stats:
            stats_data.append(["처리 시간 통계", "", ""])
            stats_data.append(["지표", "값(초)", ""])
            stats_data.append(["평균", f"{sum(processing_time_stats)/len(processing_time_stats):.2f}", ""])
            stats_data.append(["최대", f"{max(processing_time_stats):.2f}", ""])
            stats_data.append(["최소", f"{min(processing_time_stats):.2f}", ""])
            stats_data.append(["총합", f"{sum(processing_time_stats):.2f}", ""])
        
        # 데이터 입력 및 스타일 적용
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 헤더 스타일 적용
                if any(header in str(value) for header in ["통계", "구분", "지표", "타입"]):
                    cell.style = "stat_header_style"
        
        # 차트 생성
        if include_charts and plate_type_stats:
            self._create_charts(ws, plate_type_stats, len(stats_data) + 3)
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_failed_analysis_sheet(self, failed_results: List[Dict[str, Any]]):
        """실패 분석 시트 생성"""
        ws = self.wb.create_sheet("❌ 실패 분석")
        
        # 헤더
        headers = ["번호", "파일명", "오류 유형", "오류 메시지", "파일 크기", "처리 시간"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.style = "header_style"
        
        # 실패 데이터 분석
        error_types = {}
        
        for idx, result in enumerate(failed_results, 2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=result.get('file_name', ''))
            
            error_msg = result.get('error', '알 수 없는 오류')
            error_type = self._categorize_error(error_msg)
            error_types[error_type] = error_types.get(error_type, 0) + 1
            
            ws.cell(row=idx, column=3, value=error_type)
            ws.cell(row=idx, column=4, value=error_msg)
            ws.cell(row=idx, column=5, value=result.get('file_size', 'N/A'))
            ws.cell(row=idx, column=6, value=f"{result.get('processing_time', 0):.2f}초")
        
        # 오류 유형별 통계 추가
        stat_start_row = len(failed_results) + 4
        ws.cell(row=stat_start_row, column=1, value="오류 유형별 통계").style = "stat_header_style"
        ws.cell(row=stat_start_row, column=2, value="건수").style = "stat_header_style"
        
        for idx, (error_type, count) in enumerate(sorted(error_types.items()), stat_start_row + 1):
            ws.cell(row=idx, column=1, value=error_type)
            ws.cell(row=idx, column=2, value=count)
        
        # 열 너비 조정
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['D'].width = 40  # 오류 메시지 열 넓게
    
    def _add_image_to_cell(self, ws, row: int, col: int, image_path: str, size: tuple):
        """셀에 이미지 추가"""
        try:
            if os.path.exists(image_path):
                img = Image(image_path)
                img.width, img.height = size
                
                # 셀 위치 계산
                cell_address = ws.cell(row=row, column=col).coordinate
                ws.add_image(img, cell_address)
                
        except Exception as e:
            logger.warning(f"이미지 추가 실패 ({image_path}): {e}")
            # 이미지 추가 실패 시 텍스트로 대체
            ws.cell(row=row, column=col, value="[이미지 로드 실패]")
    
    def _create_charts(self, ws, plate_type_stats: Dict[str, int], start_row: int):
        """차트 생성"""
        try:
            # 파이 차트 생성
            pie_chart = PieChart()
            pie_chart.title = "번호판 타입별 분포"
            
            # 차트용 데이터 생성
            chart_data_start = start_row
            ws.cell(row=chart_data_start, column=5, value="번호판 타입")
            ws.cell(row=chart_data_start, column=6, value="개수")
            
            for idx, (plate_type, count) in enumerate(sorted(plate_type_stats.items()), chart_data_start + 1):
                ws.cell(row=idx, column=5, value=plate_type)
                ws.cell(row=idx, column=6, value=count)
            
            # 데이터 참조 설정
            data = Reference(ws, min_col=6, min_row=chart_data_start, max_row=chart_data_start + len(plate_type_stats))
            categories = Reference(ws, min_col=5, min_row=chart_data_start + 1, max_row=chart_data_start + len(plate_type_stats))
            
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(categories)
            
            # 차트 위치 설정
            ws.add_chart(pie_chart, f"H{chart_data_start}")
            
        except Exception as e:
            logger.warning(f"차트 생성 실패: {e}")
    
    def _categorize_error(self, error_message: str) -> str:
        """오류 메시지를 카테고리로 분류"""
        error_msg_lower = error_message.lower()
        
        if any(keyword in error_msg_lower for keyword in ['file not found', '파일', 'path']):
            return "파일 오류"
        elif any(keyword in error_msg_lower for keyword in ['image', '이미지', 'decode', 'format']):
            return "이미지 형식 오류"
        elif any(keyword in error_msg_lower for keyword in ['memory', '메모리', 'out of memory']):
            return "메모리 오류"
        elif any(keyword in error_msg_lower for keyword in ['timeout', '시간', 'time']):
            return "처리 시간 초과"
        elif any(keyword in error_msg_lower for keyword in ['model', '모델', 'inference']):
            return "모델 추론 오류"
        else:
            return "기타 오류"
    
    def export_batch_results(
        self, 
        batch_summary, 
        detailed_results: List[Dict], 
        output_path: str,
        **kwargs
    ) -> str:
        """배치 처리 결과를 엑셀로 출력
        
        Args:
            batch_summary: BatchProcessSummary 객체
            detailed_results: 상세 결과 리스트
            output_path: 출력 파일 경로
            **kwargs: 추가 옵션들
            
        Returns:
            생성된 파일 경로
        """
        
        # 배치 요약 정보를 딕셔너리로 변환
        summary_dict = {
            'total_files': batch_summary.total_files,
            'processed_files': batch_summary.processed_files,
            'success_rate': batch_summary.success_rate_percent,
            'total_plates': batch_summary.total_plates_detected,
            'valid_plates': batch_summary.total_valid_plates,
            'processing_time': getattr(batch_summary, 'total_processing_time', 0),
            'throughput': batch_summary.throughput_files_per_min,
            'plate_types': batch_summary.plate_type_distribution or {}
        }
        
        # 상세 결과에 배치 요약 정보 추가
        enhanced_results = []
        for result in detailed_results:
            enhanced_result = result.copy()
            enhanced_result['batch_summary'] = summary_dict
            enhanced_results.append(enhanced_result)
        
        return self.create_comprehensive_report(
            enhanced_results,
            output_path,
            **kwargs
        )


def create_sample_excel_report():
    """샘플 엑셀 보고서 생성 (테스트용)"""
    
    sample_results = [
        {
            'file_name': 'car1.jpg',
            'success': True,
            'processing_time': 1.23,
            'plates': [
                {
                    'text': '12가3456',
                    'confidence': 0.95,
                    'type': '자가용(신형)',
                    'color': '흰색',
                    'is_valid': True
                }
            ]
        },
        {
            'file_name': 'car2.jpg', 
            'success': True,
            'processing_time': 0.87,
            'plates': [
                {
                    'text': '34나5678',
                    'confidence': 0.82,
                    'type': '자가용(구형)',
                    'color': '노란색',
                    'is_valid': True
                }
            ]
        },
        {
            'file_name': 'car3.jpg',
            'success': False,
            'error': 'Image decode error',
            'processing_time': 0.12,
            'plates': []
        }
    ]
    
    exporter = ExcelExporter()
    output_file = 'sample_report.xlsx'
    
    return exporter.create_comprehensive_report(
        sample_results, 
        output_file,
        include_images=False,  # 테스트용으로 이미지 제외
        include_statistics=True,
        include_charts=True
    )


if __name__ == "__main__":
    # 테스트 실행
    try:
        output_file = create_sample_excel_report()
        print(f"샘플 엑셀 보고서가 생성되었습니다: {output_file}")
    except Exception as e:
        print(f"오류 발생: {e}")