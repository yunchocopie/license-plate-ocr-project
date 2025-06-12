from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import os


def append_to_excel(data_list, excel_path):
    """
    data_list: [(사진경로, 차량번호1, 차량번호2), ...]
    """
    if os.path.exists(excel_path): # 기존 엑셀 파일이 있으면 이어서 쓰기
        wb = load_workbook(excel_path)
        ws = wb.active
    else: # 없다면 새로 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "차량 정보"
        ws.append(["연번", "사진", "차량번호1", "차량번호2"])  # 헤더

    start_row = ws.max_row + 1

    for idx, (img_path, num1, num2) in enumerate(data_list, start=start_row):
        ws.cell(row=idx, column=1, value=idx - 1)  # 연번
        ws.cell(row=idx, column=3, value=num1)
        ws.cell(row=idx, column=4, value=num2)

        try:
            img = Image(img_path)
            img.width = 80
            img.height = 60
            ws.add_image(img, f'B{idx}')  # 사진은 B열

            ws.row_dimensions[idx].height = 45
            ws.column_dimensions['B'].width = 15
        except Exception as e:
            print(f"이미지 추가 실패 ({img_path}): {e}")

    header_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")  # 1행
    index_fill = PatternFill(fill_type="solid", fgColor="FFFFCC")   # 1열

    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).fill = header_fill

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = index_fill

    """
    글자 중앙 정렬
    """
    center_align = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align

    # 필드 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(excel_path)
